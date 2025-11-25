"""
database_functions provides functionalities for connection to the database and query from it in the context of the wwtwin package.
Copyright (C) 2025  Saba Daneshgar, BIOMATH, Ghent University

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""

from __future__ import annotations

import sys
import os
from os import listdir
import pandas as pd
# import scipy as sp
import numpy as np
# import matplotlib.pyplot as plt   
# import xlrd
import sqlite3
from sqlite3 import Error
# import pyodbc
# import psycopg2
# import pymysql
from sqlalchemy import create_engine
from sqlalchemy import text
import openpyxl
from typing import List, Sequence, Union, Optional, Any, Dict, Literal
import logging
from pathlib import Path
from sqlalchemy.engine import Engine, Connection
from urllib.parse import quote_plus

logger = logging.getLogger(__name__)
SqlConn = Union[Engine, Connection, Any]  # Any allows a DB-API connection as a fallback


def _ensure_linked_server(
    engine: Engine,
    *,
    linked_server_name: str,
    linked_server_provider: str = "Historian OLE DB Provider",
    linked_server_datasource: Optional[str] = None,
) -> None:
    """
    Ensure a linked server exists on the connected MSSQL instance.

    - No-op if the linked server already exists (checks sys.servers).
    - Creates the linked server if missing.
    """
    if not linked_server_name:
        raise ValueError("linked_server_name must be provided.")

    # Check existence
    exists_sql = text(
        """
        SELECT 1
        FROM sys.servers
        WHERE name = :srvname
        """
    )

    # Build EXEC for sp_addlinkedserver. We can't parameterize identifiers directly,
    # so we use quoted string parameters for provider and datasrc values.
    create_sql = text(
        """
        EXEC master.dbo.sp_addlinkedserver
            @server = :srvname,
            @provider = :provider,
            @datasrc = :datasrc
        """
    )

    with engine.begin() as conn:
        res = conn.execute(exists_sql, {"srvname": linked_server_name}).fetchone()
        if res:
            logger.info("Linked server '%s' already exists.", linked_server_name)
            return

        ds = linked_server_datasource or ""
        conn.execute(
            create_sql,
            {
                "srvname": linked_server_name,
                "provider": linked_server_provider,
                "datasrc": ds,
            },
        )
        logger.info(
            "Created linked server '%s' (provider='%s', datasource='%s').",
            linked_server_name,
            linked_server_provider,
            ds,
        )






def create_database_engine(
    database_type: str,
    database_name: str,
    path_to_database: Optional[Union[str, Path]] = None,
    *,
    server: Optional[str] = None,
    user: Optional[str] = None,
    password: Optional[str] = None,
    port: Optional[int] = None,
    driver: Optional[str] = None,
    readonly: bool = True,
    mode: str = "ro",
    # Historian / linked-server specifics:
    linked_server_name: str = "historian_db",
    linked_server_provider: str = "Historian OLE DB Provider",
    linked_server_datasource: Optional[str] = None,
    **kwargs: Any,
) -> Engine:
    """
    Create a SQLAlchemy Engine for the given database.
    Supports: sqlite, postgres, mysql, mssql, historian (MSSQL + linked server).
    """
    dbt = (database_type or "").strip().lower()
    if dbt not in {"sqlite", "postgres", "mysql", "mssql", "historian"}:
        raise ValueError('database_type must be one of: "sqlite", "postgres", "mysql", "mssql", "historian".')

    server = server or "localhost"
    user = user or ""
    password = password or ""
    connect_args: dict[str, Any] = {}
    url: str

    if dbt == "sqlite":
        base = Path(path_to_database or ".").expanduser().resolve()
        base.mkdir(parents=True, exist_ok=True)
        db_path = base / database_name
        uri = f"{db_path.as_posix()}?mode={mode}&cache=shared"
        url = f"sqlite+pysqlite:///{uri}"
        connect_args["uri"] = True

    elif dbt == "postgres":
        port = port or 5432
        if readonly:
            connect_args["options"] = "-c default_transaction_read_only=on"
        url = f"postgresql+psycopg://{quote_plus(user)}:{quote_plus(password)}@{server}:{port}/{database_name}"

    elif dbt == "mysql":
        port = port or 3306
        url = f"mysql+pymysql://{quote_plus(user)}:{quote_plus(password)}@{server}:{port}/{database_name}"

    elif dbt in {"mssql", "historian"}:
        port = port or 1433
        driver = driver or "ODBC Driver 17 for SQL Server"
        odbc_params = f"driver={driver.replace(' ', '+')}"
        if readonly:
            odbc_params += "&ApplicationIntent=ReadOnly"
        odbc_params += "&TrustServerCertificate=yes"
        url = (
            f"mssql+pyodbc://{quote_plus(user)}:{quote_plus(password)}@"
            f"{server}:{port}/{database_name}?{odbc_params}"
        )

    else:
        raise ValueError(f"Unsupported database_type: {database_type}")

    engine = create_engine(url, future=True, connect_args=connect_args)
    logger.info("Created SQLAlchemy engine for %s (%s)", database_name, dbt)

    # Historian path: ensure linked server exists
    if dbt == "historian":
        _ensure_linked_server(
            engine,
            linked_server_name=linked_server_name,
            linked_server_provider=linked_server_provider,
            linked_server_datasource=linked_server_datasource,
        )

    return engine












def read_hist_table(engine: Engine, linked_server: str, remote_query: str):
    """
    Example: remote_query might be something like:
      SELECT TagName, DateTime, Value FROM ihRawData WHERE TagName IN ('P1','P2') AND DateTime >= '2024-01-01'
    """
    sql = text(f"SELECT * FROM OPENQUERY([{linked_server}], :q)")
    # Note: some drivers require the remote query string to be single-quoted in T-SQL;
    # parameter binding helps avoid quoting issues.
    with engine.connect() as conn:
        return pd.read_sql_query(sql, conn, params={"q": remote_query})
    


def create_database_connection(
    database_type: str,
    database_name: str,
    path_to_database: Optional[Union[str, Path]],
    readonly: bool = True,
    mode: str = "ro",
    **kwargs: Any,
) -> Engine:
    """
    Backward-compatible shim that now returns a SQLAlchemy Engine.
    If legacy code needs a DB-API connection, use:
        engine = create_database_connection(...);
        conn = engine.raw_connection()     # DB-API connection
    """
    eng = create_database_engine(
        database_type=database_type,
        database_name=database_name,
        path_to_database=path_to_database,
        readonly=readonly,
        mode=mode,
        **kwargs,
    )

    return eng

# def create_database_connection(database_type, database_name, path_to_database, readonly=True, mode='ro', **kwargs):
#     """
#     creates a connection to a database 
    

#     Parameters
#     ----------
#     database_type : str
#         type of the database, possible values: "mysql", "postgres", "sqlite", "mssql", "historian"
#     database_name : str
#         name of the database (name of the database file in case of sqlite)
#     user : str
#         username for authentication
#     password : str
#         password for authentication
#     server : str
#         server address/name of the database
#     port : int
#         port number for the database
#     path_to_db : str
#         path to the database file, only valid for the sqlite
#     driver : str
#         driver for database connection
    
#     Returns
#     -------
#     a connection to the specified database
#     """
    
#     if database_type == 'mysql':
#         connection = create_mysql_connection(kwargs.get('server', 'localhost'), database_name, 
#                                        kwargs.get('user'), kwargs.get('password'), 
#                                        kwargs.get('port'), kwargs.get('driver'), readonly)
#     elif database_type == 'postgres':
#         connection = create_postgres_connection(kwargs.get('server', 'localhost'), database_name, 
#                                           kwargs.get('user'), kwargs.get('password'), 
#                                           kwargs.get('port'), kwargs.get('driver'), readonly)
#     elif database_type == 'sqlite':
#         if path_to_database:
#             connection = create_sqlite_connection(path_to_database, database_name, mode)
#         else:
#             raise ValueError('path_to_database is not defined!')
#     elif database_type == 'historian':
#         connection = create_hist_connection(kwargs.get('server', 'localhost'), database_name, 
#                                       kwargs.get('user'), kwargs.get('password'), 
#                                       kwargs.get('port'), kwargs.get('driver'), 
#                                       kwargs.get('linked_server_name'), kwargs.get('linked_server_provider'),
#                                       kwargs.get('linked_server_datasource'), readonly)
#     elif database_type == 'mssql':
#         connection = create_mssql_connection(kwargs.get('server', 'localhost'), database_name, 
#                                        kwargs.get('user'), kwargs.get('password'), 
#                                        kwargs.get('port'), kwargs.get('driver'), readonly)
#     else:
#         raise ValueError('Use a valid database type, possible options: "mysql", "postgres", "sqlite", "historian" and "mssql"')

#     return connection

    
# def create_mssql_connection(server, database_name, user, password, port, driver=None, readonly=True):
#     """
#     creates a connection to a Microsoft SQL Server database

#     Parameters
#     ----------
#     server: str
#         name of the host server
#     database_name: 
#         database name
#     user: str
#         username for authentication
#     password: str
#         password for authentication
#     port: int
#         connection port number
#     driver: str
#         the driver for Microsoft SQL Server, will be set automatically if is None

    
#     Returns
#     -------
#     connection to the Microsoft SQL Server database

#     """

#     connection = None
#     if driver is None:
#         driver = '{ODBC Driver 17 for SQL Server}'


#     connection_string = (
#     f'DRIVER={driver};'
#     f'SERVER={server};'
#     f'DATABASE={database_name};'
#     f'UID={user};'
#     f'PWD={password};'
#     f'PORT={port};'
#     f'charset=utf8mb4;'
#     )


    
#     try:
#         connection = pyodbc.connect(connection_string, readonly)
#     except Error:
#         print('Connection to database failed!')

#     return connection
    



# def create_hist_connection(server, database_name, user, password, port, driver=None, 
#                            linked_server_name='historian_db', 
#                            linked_server_provider='Historian OLE DB Provider',
#                            linked_server_datasource=None, readonly=True):
#     """
#     creates a linked server on a Microsoft SQL Server as an interface to the historian database
#     creates a connection to the Microsoft SQL Server

#     Parameters
#     ----------
#     server: str
#         name of the host server, this should be a Microsoft SQL Server
#     database_name: 
#         database name
#     user: str
#         username for authentication
#     password: str
#         password for authentication
#     port: int
#         connection port number
#     driver: str
#         the driver for Microsoft SQL Server, will be set automatically if is None
#     linked_server_name: str
#         the name of the linked server to create for the historian database
#     linked_server_provide: str
#         name of the specific OLE DB provider that corresponds to the historian data source
#     linked_server_datasource: str
#         name of the historian server (data source)

    
#     Returns
#     -------
#     connection to the historian database
    
#     """
    
    

#     connection = create_mssql_connection(server, database_name, user, password, port, driver, readonly)
#     cursor = connection.cursor()
#     create_linkedserver_cmd = f"EXEC master.dbo.sp_addlinkedserver @server = N'{linked_server_name}', @provider = N'{linked_server_provider}', @datasrc = N{linked_server_datasource}'"
#     cursor.execute(create_linkedserver_cmd)
    
#     return connection

# def create_mysql_connection(server, database_name, user, password, port, driver=None, readonly=True):
#     """
#     creates a connection to a MySQL database

#     Parameters
#     ----------
#     server: str
#         name of the host server
#     database_name: 
#         database name
#     user: str
#         username for authentication
#     password: str
#         password for authentication
#     port: int
#         connection port number
#     driver: str
#         the driver for MySQL, will be set automatically if is None

    
#     Returns
#     -------
#     connection to the MySQL database

#     """

#     connection = None
#     if driver is None:
#         driver = '{MySQL ODBC 8.0 ANSI Driver}'


#     connection_string = (
#     f'DRIVER={driver};'
#     f'SERVER={server};'
#     f'DATABASE={database_name};'
#     f'UID={user};'
#     f'PWD={password};'
#     f'PORT={port};'
#     f'charset=utf8mb4;'
#     )


    
#     try:
#         connection = pyodbc.connect(connection_string, readonly)
#     except Error:
#         print('Connection to database failed!')

#     return connection


# def create_postgres_connection(server, database_name, user, password, port, driver=None, readonly=True):
#     """
#     creates a connection to a PostgreSQL database

#     Parameters
#     ----------
#     server: str
#         name of the host server
#     database_name: 
#         database name
#     user: str
#         username for authentication
#     password: str
#         password for authentication
#     port: int
#         connection port number
#     driver: str
#         the driver for PostgreSQL, will be set automatically if is None

    
#     Returns
#     -------
#     connection to the PostgreSQL database

#     """
#     connection = None
#     if driver is None:
#         driver = '{PostgreSQL Unicode}'


#     connection_string = (
#     f'DRIVER={driver};'
#     f'SERVER={server};'
#     f'DATABASE={database_name};'
#     f'UID={user};'
#     f'PWD={password};'
#     f'PORT={port};'
#     )

#     try:
#         connection = pyodbc.connect(connection_string, readonly)
#     except Error as err:
#         print('Connection to database failed!')
    
#     return connection

# def create_sqlite_connection(path_to_db, database_name, mode='ro'):
#     """
#     creates a connection to the SQLite3 local database

#     Parameters
#     ----------
#     path_to_db : str
#         path to the database file (with extension .db)
#     database_name: str
#         name of the database file without the extension
    
#     Returns
#     -------
#     a connection to the SQLite3 database
#     """

#     connection = None
#     try:
        
#         from urllib.request import pathname2url

#         db_name = path_to_db +  database_name + '.db'
#         db_uri = f'file:{pathname2url(db_name)}?mode={mode}'
#         # print(path_to_db + database_name + '.db')
#         connection = sqlite3.connect(db_uri, uri=True)
        
#     except Error as er:
#         print('Connection to database failed!')

#     return connection






# def write_to_db(data, table_name, connection, method='replace', index=False, index_label=None):
#     """
#     writes data to a table in a database

#     Parameters
#     ----------
#     data : DataFrame
#         a pandas dataframe that contains the data
#     table_name : str
#         name of the table to be created in the database for the data
#     connection : a database connection
#         a connection to the database
#     method : str
#         the method to be used if the table already exists in the database. options: 'fail', 'replace', 'append'
#     index : bool
#         if True, writes dataframe index as a column in the database table
#     index_label : str
#         the name of the index column 

#     Returns
#     -------
#     None
#     """
#     data.to_sql(table_name, connection, if_exists=method, index=index, index_label=index_label)

#     return None


# def query_from_db(sql_command, connection, index_column, date_column_name):
#     """
#     makes a request to query data from a table in the database

#     Parameters
#     ----------
#     sql_command : str
#         a SQL query command for a database
#     connection : database connection
#         a connection to the sqlite3 database
#     index_column : str
#         column to be set as the index
#     date_column_name : list
#         list of column names to parse as dates


#     Returns
#     -------
#     a pandas dataframe that contains the data
#     """
#     connection.begin()
#     data = pd.read_sql_query(text(sql_command), connection, index_col=index_column, parse_dates=date_column_name)
#     connection.close()
#     return data


# def read_database_data(table_name, tagname_dict, historian, time_column_name, **kwargs):
#     """
#     reads data from a database

#     Parameters
#     ----------
#     table_name: str
#         name of the table in the database to read the data from
#     tagname_dict: dict
#         a dictionary containing the list of tag names for the columns of data to read
#     historian: str
#         name of the historian database (only if the database type is historian)
#     time_column_name: str
#         name of the column header where the time data is stored

#     Returns
#     -------
#     a pandas dataframe that contains the data

#     """
    
    
#     database_config = {**kwargs}

#     if all(config in database_config.keys() for config in ('database_type' and 'database_name')) == False:
#         raise ValueError('At least "database_type" and "database_name" should be defined for reading data from a database!') 

#     conn = create_database_connection(database_type=database_config.get('database_type'), database_name=database_config.get('database_name'), 
#                                       user=database_config.get('user'), password=database_config.get('password'), server=database_config.get('server'),
#                                       port=database_config.get('port'), path_to_db=database_config.get('path_to_db'), driver=database_config.get('driver'),
#                                       linked_server_name=database_config.get('linked_server_name'), linked_server_provider=database_config.get('linked_server_provider'),
#                                       linked_server_datasource=database_config.get('linked_server_datasource'))

#     df = create_dataframe(table_name, tagname_dict, conn, historian, time_column_name)
    
#     return df


# def create_query_cmd(table_name, historian=None, time_column_name=None, variable_column_names=None, select_all=True, **kwargs):
#     """
#     create a command of SQL query for the ihistorian database

#     Parameters
#     ----------
#     table_name : str
#         name of the table in the ihistorian database
#     historian: str
#         name of the historian database to query from
#     time_column_name: str
#         name of the column header where time data is stored
#     **kwargs
#         the keyword arguments are used for filtering the select command in sql
    
#     Returns
#     -------
#     the SQL query command 

#     """

#     sql_cmd = list()
    
    
#     if historian:
#         if select_all:
#             sql_cmd.append(f"SELECT * FROM OPENQUERY({historian}, 'SELECT * FROM {table_name}")
#         else:
#             allcolumns = ', '.join(variable_column_names)
#             sql_cmd.append(f"SELECT * FROM OPENQUERY({historian}, 'SELECT {time_column_name}, {allcolumns} FROM {table_name}")
            
#     else:
#         if select_all:
#             sql_cmd.append(f"SELECT * FROM {table_name}")
#         else:
#             allcolumns = ', '.join(variable_column_names)
#             sql_cmd.append(f"SELECT {time_column_name}, {allcolumns} FROM {table_name}")

#     if kwargs:
#         sql_cmd.append(" WHERE ")
#         for k, v in kwargs.items():
#             if k == 'starttime':
#                 sql_cmd.append(f"{time_column_name} >= {v} AND ")
#             elif k == 'endtime':
#                 sql_cmd.append(f"{time_column_name} <= {v} AND ")
#             # else:
#             #     sql_cmd.append(f"{k} = {v} AND ")
            
#             elif k == 'filter_eq':
#                 for _k, _v in v.items():
#                     sql_cmd.append(f"{_k} = {_v} AND ")  
#             elif k == 'filter_gt':
#                 for _k, _v in v.items():
#                     sql_cmd.append(f"{_k} > {_v} AND ")
#             elif k == 'filter_geg':
#                 for _k, _v in v.items():
#                     sql_cmd.append(f"{_k} >= {_v} AND ")
#             elif k == 'filter_lt':
#                 for _k, _v in v.items():
#                     sql_cmd.append(f"{_k} < {_v} AND ")
#             elif k == 'filter_leq':
#                 for _k, _v in v.items():
#                     sql_cmd.append(f"{_k} <= {_v} AND ")
#             else:
#                 sql_cmd.append(f"{k} = {v}")

            
#         if historian:
#             sql_cmd.append("')")
#     elif historian:
#             sql_cmd.append("')")
        
#     full_cmd = "".join(sql_cmd)
#     if full_cmd.endswith('AND '):
#         full_cmd = full_cmd[:-5]

#     return full_cmd
 
    
# def create_query_cmd_list(table_name, historian, time_column_name, variable_column_names, select_all, starttime, endtime, 
#                           filter_tagnames_column, filter_tagnames, filter_eq, filter_gt, filter_lt, filter_geq, filter_leq):
#     """
#     create a list of SQL query commands for multiple parameters

#     Parameters
#     ----------
#     table_name : str
#         name of the table to query the data in the ihistorian database
#     tagname_dict : dict
#         a dictionary that contains tagnames for the columns to query the data
    
#     Returns
#     -------
#     a list of SQL query command

#     """
#     sql_cmd_list = []
#     query_options = {}
#     query_config_list = []
#     if starttime:
#         query_options['starttime'] = starttime
    
#     if endtime:
#         query_options['endtime'] = endtime

#     # if filter_tagnames_column:
#     #     query_options['filter_tagnames_column'] = filter_tagnames_column
    
#     # if filter_eq:
#     #     if filter_tagnames_column:
#     #         if isinstance(filter_tagnames_column, list):
#     #             assert len(filter_tagnames_column)==len(filter_eq), 'The length of "filter_tagnames_column" and "filter_eq" should be the same!'
#     #             for n,v in zip(filter_tagnames_column, filter_eq):
#     #                 query_options[n] = v
                    
#     #         elif isinstance(filter_tagnames_column, str):
#     #                 query_options[filter_tagnames_column] = filter_eq
                    
#     #         else:
#     #             raise TypeError('Invalid "filter_tagnames_column"! It should be either a string or a list of strings.')
#     #         query_config_list.append(query_options)
#     #         # for n, v in map(lambda c: (filter_tagnames_column, c), filter_eq):
#     #         #     query_options[n] = v
#     #         #     query_config_list.append(query_options)
#     #         # query_options['filter_eq'] = filter_eq
#     #     else:
#     #         raise ValueError('To use "filter_eq", "filter_tagnames", "filter_gt", "filter_lt", "filter_geq" and "filter_leq" options, the variable "filter_tagnames_column" should be defined.')
    
#     if filter_eq:
#         query_options['filter_eq'] = filter_eq
#     if filter_gt:
#         query_options['filter_gt'] = filter_gt
#     if filter_geq:
#         query_options['filter_geq'] = filter_geq
#     if filter_lt:
#         query_options['filter_lt'] = filter_lt
#     if filter_leq:
#         query_options['filter_leq'] = filter_leq
        
#     # query_config_list.append(query_options)
    
#     if filter_tagnames_column is not None and filter_tagnames is not None:
#         if isinstance(filter_tagnames_column, str):
#             for n, v in map(lambda c: (filter_tagnames_column, c), filter_tagnames):
#                 query_options_temp = query_options.copy()
#                 # query_config_list.append({n : v})
#                 query_options_temp[n] = v
#                 query_config_list.append(query_options_temp)
#     else:
#         query_config_list.append(query_options)

        
#     # if filter_tagnames:
#     #     if filter_tagnames_column:
#     #         if isinstance(filter_tagnames, list):
#     #             pass
#     #         elif isinstance(filter_tagnames, str):
#     #             pass
#     #         else:
#     #             raise TypeError('Invalid "filter_tagnames"! It should be either a string or a list of strings.')

#     #     else:
#     #         raise ValueError('To use "filter_eq", "filter_tagnames", "filter_gt", "filter_lt", "filter_geq" and "filter_leq" options, the variable "filter_tagnames_column" should be defined.')




#     print(query_config_list)
    
#     if len(query_config_list) == 0:
#         sql_cmd = create_query_cmd(table_name, historian, time_column_name, variable_column_names, select_all)
#         sql_cmd_list.append(sql_cmd)
    
#     else:
#         for q in query_config_list:
#             sql_cmd = create_query_cmd(table_name, historian, time_column_name, variable_column_names, select_all, **q)

#             sql_cmd_list.append(sql_cmd)

#     ### TO DO: add different filters in a single dict and pass it to query cmd
#     # filter_tagnames_column, filter_tagnames, filter_eq, filter_gt, filter_lt, filter_geq, filter_leq
#     # if filter_tagnames_column and filter_eq:
#     #     for n, v in map(lambda c: (filter_tagnames_column, c), filter_eq):
#     #         query_config_list.append({'starttime' : starttime, 'endtime' : endtime, n : v})

    

#     # else:
#     #     for q in query_config_list:
#     #         q_ = {k:v for k,v in q.items() if v is not None}
#     #         sql_cmd = create_query_cmd(table_name, historian, time_column_name, variable_column_names, select_all, **q_)
#     #         sql_cmd_list.append(sql_cmd)
                             
#     # for k, v in tagname_dict.items():
#     #    sql_cmd = create_query_cmd(table_name, historian, time_column_name, k=v)
#     #    sql_cmd_list.append(sql_cmd)
    
#     return sql_cmd_list


# def create_dataframe(conn, table_name, historian, time_column_name, variable_column_names, select_all, starttime, endtime, 
#                      filter_tagnames_column, filter_tagnames, filter_eq, filter_gt, filter_lt, filter_geq, filter_leq):
#     """
#     creates a dataframe that contains the raw data from ihistorian database

#     Parameters
#     ----------
#     sql_cmd_list : list
#         a list of SQL query command for multiple columns
#     tagname_dict : dict
#         a dictionary that contains tagnames for the columns to query the data
#     conn : database connection
#         a connection to the ihistorian database
    
#     Returns
#     -------
#     a pandas dataframe that contains the data from the ihistorian database

#     """

#     sql_cmd_list = create_query_cmd_list(table_name, historian, time_column_name, variable_column_names, select_all, starttime, endtime, filter_tagnames_column, filter_eq
#                                          ,filter_tagnames, filter_gt, filter_lt, filter_geq, filter_leq)
#     print(sql_cmd_list)
#     # if select_all:
#     #     df = pd.read_sql(sql_cmd_list[0], conn)
        
#     # else:
#     #     df = pd.read_sql(sql_cmd_list[0], conn)
#     #     # df = df[[time_column_name,variable_column_names]]
    
#     df = pd.read_sql(sql_cmd_list[0], conn)

#     for i in range(1, len(sql_cmd_list)):
#         df_ = pd.read_sql(sql_cmd_list[i], conn)
#         # df = pd.concat([df,df_[[variable_column_names]]], axis=1)
#         df = pd.concat([df,df_], axis=1)

#     # column_names = list(tagname_dict.keys())
#     # if filter_tagnames_column and filter_eq:
#     #     if isinstance(filter_eq, str):
#     #         column_names = [filter_eq]
#     #     else:    
#     #         column_names = filter_eq

#     #     column_names.insert(0, 'Datetime')
#     #     df.columns = column_names
    
#     return df




def _as_list(x: Optional[Union[str, Sequence[str]]]) -> Optional[List[str]]:
    if x is None:
        return None
    return [x] if isinstance(x, str) else list(x)


def _get_sqlalchemy_connection(conn: SqlConn) -> Connection:
    """
    Accept an Engine, Connection, or DB-API connection and return a SQLAlchemy Connection.
    Closes the temporary wrapper automatically at the end of the calling context.
    """
    if isinstance(conn, Connection):
        return conn
    if isinstance(conn, Engine):
        return conn.connect()
    # DB-API fallback: wrap via SQLAlchemy's creator pattern (one-shot)
    from sqlalchemy import create_engine

    raw = conn  # DB-API connection object
    engine = create_engine("sqlite://", creator=lambda: raw)  # URL string is ignored by creator
    return engine.connect()


def create_dataframe(
    conn: SqlConn,
    table_name: Optional[str],
    is_historian: bool,
    time_column: Optional[str],
    variable_columns: Optional[Union[str, Sequence[str]]],
    select_all: bool = True,
    starttime: Optional[Union[str, pd.Timestamp]] = None,
    endtime: Optional[Union[str, pd.Timestamp]] = None,
    filter_tagnames_column: Optional[str] = None,
    filter_eq: Optional[Dict[str, Any]] = None,
    filter_tagnames: Optional[Union[str, Sequence[str]]] = None,
    filter_gt: Optional[Dict[str, Any]] = None,
    filter_lt: Optional[Dict[str, Any]] = None,
    filter_geq: Optional[Dict[str, Any]] = None,
    filter_leq: Optional[Dict[str, Any]] = None,
    *,
    schema: Optional[str] = None,
    limit: Optional[int] = None,
    # Historian OPENQUERY path:
    linked_server_name: Optional[str] = None,
    openquery_sql: Optional[str] = None,
) -> pd.DataFrame:
    """
    Read from a database using an Engine/Connection (or DB-API connection).

    If `is_historian` is True and `openquery_sql` & `linked_server_name` are provided,
    the function will read via:
        SELECT * FROM OPENQUERY([linked_server_name], '<openquery_sql>')
    Otherwise it will build a portable SELECT on `schema.table_name`.

    Returns
    -------
    pd.DataFrame
    """
    # Historian (OPENQUERY) branch
    if is_historian:
        if not linked_server_name or not openquery_sql:
            raise ValueError(
                "Historian mode requires both 'linked_server_name' and 'openquery_sql'."
            )
        # OPENQUERY string literal must be embedded; parameterization is driver-dependent.
        # Safely single-quote the remote SQL:
        remote_sql = openquery_sql.replace("'", "''")
        tsql = f"SELECT * FROM OPENQUERY([{linked_server_name}], '{remote_sql}')"
        with _get_sqlalchemy_connection(conn) as sa_conn:
            df = pd.read_sql_query(text(tsql), sa_conn)
            logger.info("Read %d rows via OPENQUERY on linked server '%s'", len(df), linked_server_name)
            return df

    # Regular RDBMS path
    if not table_name and not select_all:
        raise ValueError("table_name must be provided when not using historian/OPENQUERY.")

    cols_sql: str
    if select_all or not variable_columns:
        cols_sql = "*"
    else:
        cols: List[str] = _as_list(variable_columns) or []
        if time_column and time_column not in cols:
            cols = [time_column] + cols
        # naive quoting (assumes simple identifiers). For mixed-case or reserved words,
        # you can wrap with double quotes or brackets based on dialect if needed.
        cols_sql = ", ".join(cols)

    fq_table = f"{schema}.{table_name}" if schema else table_name

    where_clauses: List[str] = []
    params: Dict[str, Any] = {}

    # Time window
    if time_column:
        if starttime is not None:
            where_clauses.append(f"{time_column} >= :starttime")
            params["starttime"] = pd.to_datetime(starttime)
        if endtime is not None:
            where_clauses.append(f"{time_column} <= :endtime")
            params["endtime"] = pd.to_datetime(endtime)

    # Tag filters
    if filter_tagnames_column and filter_tagnames is not None:
        tags = _as_list(filter_tagnames) or []
        if tags:
            placeholders = ", ".join([f":tag{i}" for i in range(len(tags))])
            where_clauses.append(f"{filter_tagnames_column} IN ({placeholders})")
            params.update({f"tag{i}": t for i, t in enumerate(tags)})

    # Column comparisons
    def _cmp(prefix: str, op: str, mapping: Optional[Dict[str, Any]]):
        if not mapping:
            return
        for i, (col, val) in enumerate(mapping.items()):
            key = f"{prefix}_{i}"
            where_clauses.append(f"{col} {op} :{key}")
            params[key] = val

    _cmp("eq", "=", filter_eq)
    _cmp("gt", ">", filter_gt)
    _cmp("lt", "<", filter_lt)
    _cmp("geq", ">=", filter_geq)
    _cmp("leq", "<=", filter_leq)

    where_sql = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    limit_sql = f" LIMIT {int(limit)}" if (limit is not None and int(limit) > 0) else ""

    sql = f"SELECT {cols_sql} FROM {fq_table}{where_sql}{limit_sql}"
    
    with _get_sqlalchemy_connection(conn) as sa_conn:
        df = pd.read_sql_query(text(sql), sa_conn, params=params)
        logger.info("Read %d rows from %s", len(df), fq_table)
        print(df.head())
        return df


def write_to_WEST():
    pass


# def write_to_WEST(df,file_normal,file_west,units,filepath=os.getcwd(),fillna=True):
#         """
#         writes a text-file that is compatible with WEST. Adds the units as
#         they are given in the 'units' argument.

#         Parameters
#         ----------
#         df : pd.DataFrame
#             the dataframe to write to WEST
#         file_normal : str
#             name of the original file to write, not yet compatible with WEST
#         file_west : str
#             name of the file that needs to be WEST compatible
#         units : array of strings
#             array containing the units for the respective columns in df
#         filepath : str
#             directory to save the files in; defaults to the current one
#         fillna : bool
#             when True, replaces nan values with 0 values (this might avoid
#             WEST problems later one).

#         Returns
#         -------
#         None; writes files
#         """
#         if fillna:
#             df = df.fillna(0)
#         df.to_csv(os.path.join(filepath,file_normal),sep='\t')

#         f = open(os.path.join(filepath,file_normal),'r')
#         columns = f.readline()
#         temp = f.read()
#         f.close()

#         f = open(os.path.join(filepath,file_west), 'w')
#         f.write('#.t' + columns)
#         unit_line = '#d\t'
#         for i in range(0,len(units)-1):
#             unit_line = unit_line + '{}\t'.format(units[i])
#         unit_line = unit_line + '{}\n'.format(units[-1])
#         f.write(unit_line)
#         f.write(temp)
#         f.close()


def find_data_freq(s):
    for idx, letter in enumerate(s, 0):
        if letter.isalpha():
            freq = int(s[:idx])
            unit = s[idx:]
            return freq, unit
        


# def _read_file(filepath,ext,skiprows=0,encoding='utf8',decimal='.'):
#     """
#     Read a file of given extension and save it as a pandas dataframe

#     Parameters
#     ----------
#     filepath : str
#         the complete path to the file to be read and saved as dataframe
#     ext : str
#         the extension (in words) of the file that needs to be read and saved
#     skiprows : int
#         number of rows to skip when reading a file

#     Returns
#     -------
#     A pandas dataframe containing the data from the given file

#     """
#     # _, _ext = os.path.splitext(filepath)

#     if ext in ['.csv', '.txt']:
#         with open(filepath, 'r') as f:
#             first_line = f.readline()
#             if '\t' in first_line:
#                 sep = '\t'
#             elif ';' in first_line:
#                 sep = ';'
#             elif ',' in first_line:
#                 sep = ','
#             else:
#                 sep=None
#         return pd.read_csv(filepath, sep=sep,skiprows=skiprows,encoding=encoding,
#                            decimal=decimal, low_memory=False,index_col=None)
#     elif ext in ['.xls', '.xlsx']:
#         return pd.read_excel(filepath,skiprows=skiprows,index_col=None)
#     else:
#         raise ValueError(f'Unsupported data file extension: {ext}. Available options: ".csv", ".txt", ".xls", and ".xlsx".')



def _read_file(
    filepath: str | Path,
    ext: str,
    skiprows: int = 0,
    encoding: str = "utf8",
    decimal: str = ".",
) -> pd.DataFrame:
    """
    Read a file of given extension into a pandas DataFrame.

    Parameters
    ----------
    filepath : str | Path
        Complete path to the file.
    ext : str
        File extension, e.g. ".csv", ".txt", ".xls", ".xlsx", ".parquet", ".feather".
    skiprows : int
        Number of rows to skip at the start of the file.
    encoding : str
        File encoding (for text/CSV).
    decimal : str
        Decimal separator (for CSV/text).

    Returns
    -------
    pd.DataFrame
        Data from the file.

    Raises
    ------
    FileNotFoundError
        If the file does not exist.
    ValueError
        If the extension is unsupported.
    """
    path = Path(filepath).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = ext.lower().strip()

    try:
        if ext in (".csv", ".txt"):
            # Auto-detect separator from the first line
            sep: Optional[str] = None
            with path.open("r", encoding=encoding, errors="ignore") as f:
                first_line = f.readline()
                if "\t" in first_line:
                    sep = "\t"
                elif ";" in first_line:
                    sep = ";"
                elif "," in first_line:
                    sep = ","

            df = pd.read_csv(
                path,
                sep=sep,
                skiprows=skiprows,
                encoding=encoding,
                decimal=decimal,
                low_memory=False,
                index_col=None,
            )

        elif ext in (".xls", ".xlsx"):
            df = pd.read_excel(path, skiprows=skiprows, index_col=None)

        elif ext == ".parquet":
            df = pd.read_parquet(path)

        elif ext == ".feather":
            df = pd.read_feather(path)

        else:
            raise ValueError(
                f"Unsupported extension: {ext}. "
                "Allowed: .csv, .txt, .xls, .xlsx, .parquet, .feather"
            )

        logger.info("Read %d rows and %d cols from %s", *df.shape, path.name)
        return df

    except Exception as e:
        logger.exception("Failed to read %s: %s", path, e)
        raise

    

# def _open_file(filepath,ext='.txt'):
#     """
#     Opens file of a given extension in readable mode

#     Parameters
#     ----------
#     filepath : str
#         the complete path to the file to be opened in read mode
#     ext : str
#         the extension (in words) of the file that needs to be opened in read
#         mode

#     Returns
#     -------
#     The opened file in read mode

#     """
#     if ext in ['.txt', '.zrx', '.csv']:
#         return open(filepath, 'r')
#     elif ext in ['.xls', '.xlsx']:
#         return openpyxl.load_workbook(filepath, read_only=True)
#     else:
#         raise ValueError(f'Unsupported data file extension: {ext}. Available options: ".csv", ".txt", ".xls", and ".xlsx".')
    
# def _get_header_length(read_file,ext='.txt',comment='#'):
#     """
#     Determines the amount of rows that are part of the header in a file that is
#     already opened and readable

#     Parameters
#     ----------
#     read_file : opened file
#         an opened file object that is readable
#     ext : str
#         the extension (in words) of the file the headerlength needs to be found
#         for
#     comment : str
#         comment symbol used in the files

#     Returns
#     -------
#     headerlength : int
#         the amount of rows that are part of the header in the read file

#     """

#     headerlength = 0
#     header_test = comment
#     counter = 0
#     if ext in ['.xls', '.xlsx'] or ext == '.zrx':
#         while header_test == comment:
#             _workbook = openpyxl.load_workbook(read_file, data_only=True, read_only=True)
#             _sheet = _workbook.worksheets[0]
#             _value = _sheet.cell(row=counter + 1, column=1).value
#             # header_test = str(read_file.sheet_by_index(0).cell_value(counter,0))[0]
#             header_test = str(_value)[0]
#             headerlength += 1
#             counter +=1

#     elif ext == '.txt':
#         while header_test == comment:
#             _f = open(read_file)
#             header_test = _f.readline()[0]
#             headerlength += 1
#             _f.close()

#     return headerlength-1
def _get_header_length(
    read_file: str | Path,
    ext: str = ".txt",
    comment: str = "#",
) -> int:
    """
    Determine the number of header rows at the top of a file.

    Parameters
    ----------
    read_file : str | Path
        Path to the file.
    ext : str, default ".txt"
        File extension. Supported: ".txt", ".csv", ".xls", ".xlsx", ".zrx".
    comment : str, default "#"
        Comment symbol indicating header rows in text/CSV files.

    Returns
    -------
    int
        Number of header rows.

    Raises
    ------
    FileNotFoundError
        If the file does not exist.
    ValueError
        If the extension is unsupported or parsing fails.
    """
    path = Path(read_file).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = ext.lower().strip()
    headerlength = 0
    counter = 0

    try:
        if ext in (".xls", ".xlsx", ".zrx"):
            # Excel-like: inspect first column of successive rows until we hit a non-comment
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            try:
                sheet = wb.worksheets[0]
                while True:
                    value = sheet.cell(row=counter + 1, column=1).value
                    if value is None:
                        break
                    s = str(value)
                    if not s or s[0] != comment:
                        break
                    headerlength += 1
                    counter += 1
            finally:
                wb.close()

        elif ext in (".txt", ".csv"):
            # Text/CSV: count leading lines that start with the comment char
            with path.open("r", encoding="utf8", errors="ignore") as f:
                while True:
                    pos = f.tell()
                    line = f.readline()
                    if not line:
                        break
                    if not line or line[0] != comment:
                        # Rewind so downstream readers can start from the first data line if needed
                        f.seek(pos)
                        break
                    headerlength += 1

        else:
            raise ValueError(
                f"Unsupported extension for header detection: {ext}. "
                "Allowed: .txt, .csv, .xls, .xlsx, .zrx"
            )

        logger.debug("Detected %d header rows in %s", headerlength, path.name)
        return max(0, headerlength)

    except Exception as e:
        logger.exception("Failed to detect header length in %s: %s", path, e)
        raise

# def read_files(path,files,sep=',',comment='#',encoding='utf8',decimal='.', to_csv=False):
def read_files(
        path: Union[str, os.PathLike],
        files: Sequence[str],
        sep: str = ",",
        comment: str = "#",
        encoding: str = "utf8",
        decimal: str = ".",
        to_csv: bool = False,
    ) -> pd.DataFrame:
    """
    Reads all files in a given directory, joins them and returns one pd.dataframe

    Parameters
    ----------
    path : str
	path to the folder that contains the files to be joined
    files : list
        list of files to be joined, must be the same extension
    ext : str
        extention of the files to read; possible: excel, text, csv
    sep : str
        the separating element (e.g. , or \t) necessary when reading csv-files
    comment : str
        comment symbol used in the files
    sort : array of bool and str
        if first element is true, apply the sort function to sort the data
        based on the tags in the column mentioned in the second element of the
        sort array

    Returns
    -------
    pd.dataframe:
        pandas dataframe containin concatenated files in the given directory
    """
    # #Initialisations
    # data = pd.DataFrame()

    

    # #Select files based on extension and sort files alphabetically to make sure
    # #they are added to each other in the correct order
    # #files = list_files(path,ext)
    # files.sort()
    # # print('Extracting data from',len(files),'files...')
    # print('Extracting data...')

    # #Read files
    # for file_name in files:
    #     dir_file_path = os.path.join(path,file_name)
    #     _, _ext = os.path.splitext(file_name)
        

    #     # with _open_file(dir_file_path,_ext) as read_file:
    #     headerlength = _get_header_length(dir_file_path,_ext,comment)
    #     data = pd.concat([data, _read_file(dir_file_path,ext=_ext,
    #                                 skiprows=headerlength,
    #                                 decimal=decimal,encoding=encoding)],
    #                         ignore_index=True)
            
    #     # print('Adding file',file_name,'to dataframe')
    # if to_csv:
    #     data.to_csv('joined_files',sep=sep)

    # return data
    base = Path(path).expanduser().resolve()
    if not base.exists() or not base.is_dir():
        raise NotADirectoryError(f"Invalid directory: {base}")

    if not files:
        logger.warning("No files provided to read in %s", base)
        return pd.DataFrame()

    # Ensure deterministic order
    files_sorted: List[str] = sorted(files)

    frames: List[pd.DataFrame] = []
    for file_name in files_sorted:
        file_path = base / file_name
        if not file_path.exists():
            logger.warning("File not found, skipping: %s", file_path)
            continue

        _, ext = os.path.splitext(file_name)
        try:
            header_len = _get_header_length(str(file_path), ext, comment)  # uses your helper
        except Exception as e:
            logger.warning("Failed to detect header length for %s: %s. Assuming 0.", file_path, e)
            header_len = 0

        try:
            df = _read_file(  # your helper (must return a DataFrame)
                str(file_path),
                ext=ext,
                skiprows=header_len,
                decimal=decimal,
                encoding=encoding,
            )
            if not isinstance(df, pd.DataFrame):
                raise TypeError(f"_read_file must return a pandas.DataFrame, got {type(df)}")
            frames.append(df)
            logger.info("Read %d rows from %s", len(df), file_path.name)
        except Exception as e:
            logger.exception("Error reading %s: %s. Skipping this file.", file_path, e)
            continue

    if not frames:
        logger.warning("No data frames were read successfully from %s.", base)
        return pd.DataFrame()

    data = pd.concat(frames, ignore_index=True, sort=False)

    if to_csv:
        out_path = base / "joined_files"
        try:
            data.to_csv(out_path, sep=sep, index=False)
            logger.info("Wrote concatenated CSV to %s", out_path)
        except Exception as e:
            logger.exception("Failed to write concatenated CSV to %s: %s", out_path, e)

    return data